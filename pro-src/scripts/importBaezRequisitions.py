#!/usr/bin/env python3
"""Parse Baez Place requisitions 1–12 from Google Drive spreadsheets → JSON for LE Pro."""

from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

DRIVE_BASE = Path.home() / "My Drive/1 Martin Dorkin/Projects/334 E 176/Requisition"
REQ5_BASE = Path.home() / "My Drive/Requisition 5"
GOOGLE_API = Path.home() / ".hermes/profiles/loaf_pod/skills/productivity/google-workspace/scripts/google_api.py"
OUT = Path(__file__).resolve().parent / "baez_requisitions_import.json"

# Accessible via office@leelectrical.us token; 4 & 5 sheets live on amra account — parsed from local PDFs.
REQ_SHEETS = {
    1: "1bkMK-L2OW4szuBiaWduFNMbMtLpqzoQfcFCIjISkD0Y",
    2: "12mxNOv9t8UXcj2tFLW5h7affJGokiHl-kl3AjAneYkM",
    3: "1dzaZ2N7xezslrVsy126yaBdrBe2trMDKFAFekAS2x6U",
    6: "10JexQ0WW9PQipD46xZ9nZ_-y1S80gGy0CHF8UhX3H_U",
    7: "192b9O0p1VOGrDq-JW93MtkRqOeg5HOH19zg7W-BR8Gc",
    8: "1r0s7WZFLoABBfl9wmcY1HLeNRtmBB1jhG9sQ6T-H-fk",
    9: "1vfmxpbsgXew0kovGYowMnCwusx768Emh47vN6HVhaNw",
    10: "1btS8-EEmELUr5v90rFr_Z989kBVjM97gFsdAdEYg_yY",
    11: "15ETZY2TwnYlVFYbqKXriokCAYCwh9a2EYYbXxgYUa8o",
    12: "1CJfquOMXkGk83m37PUpyqwFy7OA5_beqiCz9DXDBNpA",
}

REQ4_FOLDER = DRIVE_BASE / "Requisition 4"
REQ4_PDF = REQ4_FOLDER / "Requisition 4 Old.pdf"
# Joy/Procore export misfiled under Requisition 4; application number is 5.
REQ5_PDF = REQ4_FOLDER / "Requestion 4 New.pdf"
PDF_REQS = {4, 5}


def pm(raw) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def norm_key(section: str, desc: str) -> str:
    sec = re.sub(r"\s+", " ", (section or "").strip().lower())
    d = re.sub(r"\s+", " ", (desc or "").strip().lower())
    return f"{sec}|{d}"


def download_xlsx(file_id: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [
            sys.executable,
            str(GOOGLE_API),
            "drive",
            "download",
            file_id,
            "--out",
            str(dest),
            "--export-mime",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ],
        check=True,
        stdout=subprocess.DEVNULL,
    )


def parse_g702(ws) -> dict:
    data: dict = {}
    for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
        txt = " ".join(str(c) for c in row if c is not None)
        if "LESS PREVIOUS CERTIFICATES" in txt:
            big = [pm(c) for c in row if pm(c) > 1000]
            if big:
                data["previousCertificates"] = max(big)
        if "CURRENT PAYMENT DUE" in txt:
            big = [pm(c) for c in row if pm(c) > 100]
            if big:
                data["currentPaymentDue"] = max(big)
        if "TOTAL COMPLETED" in txt and "STORED" in txt:
            big = [pm(c) for c in row if pm(c) > 10000]
            if big:
                data["totalCompleted"] = max(big)
        for c in row:
            if hasattr(c, "year"):
                data["periodTo"] = c.strftime("%Y-%m-%d")
    return data


def parse_g703(ws) -> list[dict]:
    rows: list[dict] = []
    section = ""
    for row in ws.iter_rows(min_row=13, max_row=300, values_only=True):
        if not any(row):
            continue
        colb, colc = row[1], row[2]
        sched = pm(row[3] if len(row) > 3 else 0)
        pct = row[8] if len(row) > 8 else 0
        desc = colc if colc else colb
        if isinstance(colb, str) and colb and sched == 0 and not colc:
            section = colb
            continue
        if not desc or sched <= 0:
            continue
        if isinstance(pct, (int, float)):
            p = round(pct * 100, 2) if pct <= 1 else round(float(pct), 2)
        else:
            p = 0.0
        rows.append(
            {
                "key": norm_key(section, str(desc)),
                "section": section,
                "description": str(desc).strip(),
                "value": sched,
                "completedPct": p,
            }
        )
    return rows


def drive_search_pdf(name: str) -> dict | None:
    try:
        out = subprocess.check_output(
            [sys.executable, str(GOOGLE_API), "drive", "search", name, "--max", "5"],
            text=True,
        )
        hits = json.loads(out)
        for h in hits:
            if h.get("name", "").lower() == name.lower():
                return {"name": h["name"], "webViewLink": h.get("webViewLink", "")}
        for h in hits:
            if name.lower().replace(" ", "") in h.get("name", "").lower().replace(" ", ""):
                return {"name": h["name"], "webViewLink": h.get("webViewLink", "")}
    except Exception:
        pass
    return None


def folder_attachments(num: int) -> list[dict]:
    folder = DRIVE_BASE / f"Requisition {num}"
    if num == 5 and not folder.exists():
        folder = REQ5_BASE
    out: list[dict] = []
    for candidate in (
        f"Requisition {num}.pdf",
        f"Req {num}.pdf",
        f"Req {num} .pdf",
    ):
        hit = drive_search_pdf(candidate)
        if hit:
            out.append(hit)
            break
    if folder.exists():
        for pdf in sorted(folder.glob("*.pdf")) + sorted(folder.glob("*.PDF")):
            if any(pdf.name == a.get("name") for a in out):
                continue
            out.append({"name": pdf.name, "localPath": str(pdf)})
    return out


def pdf_to_text(path: Path) -> str:
    dest = Path("/tmp") / f"baez_{path.stem}.txt"
    subprocess.run(
        ["pdftotext", str(path), str(dest)],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    return dest.read_text(encoding="utf-8", errors="replace")


def _next_nonempty(lines: list[str], start: int) -> int:
    j = start
    while j < len(lines) and not lines[j].strip():
        j += 1
    return j


def parse_aia_g703_text(text: str) -> dict[str, dict]:
    """Parse AIA G703 continuation sheet (Requisition 4 Old.pdf)."""
    lines = [l.strip() for l in text.split("\n")]
    for i, l in enumerate(lines):
        if l == "CONTINUATION SHEET":
            lines = lines[i:]
            break
    section = ""
    items: dict[str, dict] = {}
    floors = (
        "Subcellar Floor",
        "Cellar Floor",
        "1st Floor",
        "2nd Floor",
        "3rd Floor",
        "4th Floor",
        "5th Floor",
        "6th Floor",
        "7th Floor",
        "8th Floor",
        "9th Floor",
        "10th Floor",
        "Roof",
        "Penthouse",
    )
    i = 0
    while i < len(lines):
        l = lines[i]
        if not re.fullmatch(r"\d+", l):
            i += 1
            continue
        j = _next_nonempty(lines, i + 1)
        if j >= len(lines):
            break
        nxt = lines[j]
        if nxt in floors or nxt.startswith("CO #") or nxt == "Allowance":
            section = nxt
            i = j + 1
            continue
        desc = nxt
        sched = None
        total = None
        pct = None
        dollars: list[float] = []
        k = j + 1
        while k < min(j + 22, len(lines)):
            s = lines[k]
            if re.fullmatch(r"\d+", s) and k > j + 2:
                break
            if s.startswith("$"):
                try:
                    dollars.append(float(s.replace("$", "").replace(",", "")))
                except ValueError:
                    pass
            m = re.search(r"\$[\d,]+\.?\d*\s+(\d+\.?\d*)%", s)
            if m:
                pct = float(m.group(1))
            if re.fullmatch(r"\d+\.?\d*%", s):
                pct = float(s.replace("%", ""))
            k += 1
        if dollars:
            sched = dollars[0]
        if len(dollars) >= 4:
            total = dollars[3]
        elif len(dollars) >= 2:
            total = dollars[-1]
        if pct is None and sched and total and 0 < total <= sched * 1.05:
            pct = round(total / sched * 100, 2)
        if desc and sched and pct is not None:
            key = norm_key(section, desc)
            items[key] = {
                "section": section,
                "description": desc,
                "value": sched,
                "completedPct": pct,
            }
        i = j + 1
    return items


def parse_procore_g703_text(text: str) -> dict[str, dict]:
    """Parse Joy/Procore G703 export (Requestion 4 New.pdf — application 5)."""
    lines = [l.strip() for l in text.split("\n")]
    section = ""
    items: dict[str, dict] = {}
    seen: set[str] = set()
    i = 0
    while i < len(lines):
        if not re.fullmatch(r"\d+", lines[i]):
            i += 1
            continue
        j = _next_nonempty(lines, i + 1)
        chunk: list[str] = []
        k = j
        while k < min(j + 25, len(lines)):
            if re.fullmatch(r"\d+", lines[k]) and k > j + 3:
                break
            if lines[k]:
                chunk.append(lines[k])
            k += 1
        desc = ""
        sched = None
        pct = None
        for c in chunk:
            c2 = c.replace("5ub", "Sub").replace("5ervice", "Service")
            if c2 in (
                "Subcellar Floor",
                "Cellar Floor",
                "1st Floor",
                "2nd Floor",
                "3rd Floor",
                "4th Floor",
                "5th Floor",
                "6th Floor",
                "7th Floor",
                "8th Floor",
                "9th Floor",
                "10th Floor",
                "Roof",
                "Penthouse",
            ) or c2.startswith("CO #"):
                section = c2
            elif c2 in (
                "Electric Service Equipment",
                "Electric Service Installation",
                "Temp Electric & Lighting",
                "Testing & Inspections",
            ):
                desc = c2
            elif any(
                c2.startswith(x)
                for x in ("Feeders", "Roughing", "Low Voltage", "Fire Alarm", "Security", "Lighting", "Finish")
            ):
                desc = c2
            elif c2.startswith("$"):
                try:
                    v = float(c2.replace("$", "").replace(",", ""))
                    if v > 0 and sched is None:
                        sched = v
                except ValueError:
                    pass
            elif re.fullmatch(r"\d+\.\d+%", c2):
                pct = float(c2.replace("%", ""))
        if desc and sched is not None and pct is not None:
            key = norm_key(section, desc)
            if key not in seen:
                seen.add(key)
                items[key] = {
                    "section": section,
                    "description": desc,
                    "value": sched,
                    "completedPct": pct,
                }
        i = k if k > j else j + 1
    return items


def g702_aia_text(text: str) -> dict:
    data: dict = {}
    for label, key in (
        ("TOTAL COMPLETED & STORED", "totalCompleted"),
        ("LESS PREVIOUS CERTIFICATES", "previousCertificates"),
        ("CURRENT PAYMENT DUE", "currentPaymentDue"),
    ):
        idx = text.find(label)
        if idx < 0:
            continue
        chunk = text[idx : idx + 300].split("\n")
        vals = [pm(c) for c in chunk if pm(c) > 100]
        if vals:
            data[key] = vals[0] if key != "totalCompleted" else max(vals)
    m = re.search(r"PERIOD TO:\s*([0-9/]+)", text)
    if m:
        parts = m.group(1).split("/")
        if len(parts) == 3:
            data["periodTo"] = f"20{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
    return data


def g702_procore_text(text: str) -> dict:
    data: dict = {}
    for label, key in (
        ("Total completed and stored to date", "totalCompleted"),
        ("Less previous certificates for payment", "previousCertificates"),
        ("Current payment due:", "currentPaymentDue"),
    ):
        idx = text.find(label)
        if idx < 0:
            continue
        chunk = text[idx : idx + 250]
        vals = [pm(x) for x in re.findall(r"\$[\d,]+\.?\d*", chunk)]
        if vals:
            data[key] = vals[0]
    m = re.search(r"PERIOD:\s*([0-9/]+)\s*-\s*([0-9/]+)", text)
    if m:
        parts = m.group(2).split("/")
        if len(parts) == 3:
            data["periodTo"] = f"20{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
    return data


def build_req4_by_key(by3: dict, by5: dict, by6: dict, by4pdf: dict) -> dict[str, dict]:
    """Merge Req 4 PDF with Req 3/5/6 so billing % never runs backward."""
    merged: dict[str, dict] = {}
    for key, it6 in by6.items():
        r3 = by3.get(key, {}).get("completedPct", 0)
        r5 = by5.get(key, {}).get("completedPct", r3)
        r6 = it6.get("completedPct", 0)
        r4p = by4pdf.get(key, {}).get("completedPct")
        if r3 > r5:
            # SOV revision on later reqs — keep Req 4 aligned with Req 5/6, not stale Req 3.
            r4 = r5
        elif r4p is None or r4p > r5:
            r4 = round(r3 + (r5 - r3) * 0.5, 2)
        else:
            r4 = r4p
        r4 = min(max(r4, r3 if r3 <= r5 else r5), r5, r6)
        merged[key] = {**(by4pdf.get(key) or by5.get(key) or by3.get(key) or it6), "completedPct": r4}
    return merged


def pdf_attachments(num: int) -> list[dict]:
    out: list[dict] = []
    if num == 4:
        for path in (REQ4_PDF, REQ4_FOLDER / "Requisition 4.pdf"):
            if path.exists():
                out.append({"name": path.name, "localPath": str(path)})
    elif num == 5:
        if REQ5_PDF.exists():
            out.append({"name": "Requisition 5.pdf", "localPath": str(REQ5_PDF)})
        folder = REQ5_BASE if REQ5_BASE.exists() else DRIVE_BASE / "Requisition 5"
        if folder.exists():
            for pdf in sorted(folder.glob("*.gsheet")):
                out.append({"name": pdf.name, "localPath": str(pdf)})
    return out


def interpolate_items(before: dict, after: dict, ratio: float) -> dict:
    keys = set(before) | set(after)
    merged = {}
    for k in keys:
        b = before.get(k, {}).get("completedPct", 0)
        a = after.get(k, {}).get("completedPct", 0)
        merged[k] = {
            **(after.get(k) or before.get(k) or {}),
            "completedPct": round(b + (a - b) * ratio, 2),
        }
    return merged


def main() -> None:
    parsed: dict[int, dict] = {}
    cache = Path("/tmp/baez_req_cache")
    cache.mkdir(exist_ok=True)

    for num, fid in REQ_SHEETS.items():
        path = cache / f"req{num}.xlsx"
        if not path.exists():
            download_xlsx(fid, path)
        wb = openpyxl.load_workbook(path, data_only=True)
        g702 = parse_g702(wb["Application & Cert"])
        items = parse_g703(wb["Continuation Sheet"])
        by_key = {it["key"]: it for it in items}
        parsed[num] = {
            "num": num,
            "g702": g702,
            "items": items,
            "byKey": by_key,
            "attachments": folder_attachments(num),
        }
        print(f"req {num}: {len(items)} lines, due ${g702.get('currentPaymentDue', 0):,.0f}", file=sys.stderr)

    pdf_parsed: dict[int, dict] = {}
    if REQ4_PDF.exists():
        t4 = pdf_to_text(REQ4_PDF)
        pdf_parsed[4] = {
            "byKey": parse_aia_g703_text(t4),
            "g702": {**g702_aia_text(t4), "source": "Requisition 4 Old.pdf"},
        }
        print(
            f"req 4 pdf: {len(pdf_parsed[4]['byKey'])} lines",
            file=sys.stderr,
        )
    if REQ5_PDF.exists():
        t5 = pdf_to_text(REQ5_PDF)
        pdf_parsed[5] = {
            "byKey": parse_procore_g703_text(t5),
            "g702": {**g702_procore_text(t5), "source": "Requestion 4 New.pdf (app #5)"},
        }
        print(
            f"req 5 pdf: {len(pdf_parsed[5]['byKey'])} lines, due ${pdf_parsed[5]['g702'].get('currentPaymentDue', 0):,.0f}",
            file=sys.stderr,
        )

    # Master SOV from req 12 (includes change orders).
    master = parsed[12]["items"]
    master_keys = [it["key"] for it in master]

    reqs_out = []
    prev_by_key: dict[str, dict] = {}

    for num in range(1, 13):
        if num == 4 and num in pdf_parsed and 5 in pdf_parsed:
            by_key = build_req4_by_key(
                parsed[3]["byKey"],
                pdf_parsed[5]["byKey"],
                parsed[6]["byKey"],
                pdf_parsed[4]["byKey"],
            )
            g702 = pdf_parsed[4]["g702"]
            attachments = pdf_attachments(num) or folder_attachments(num)
        elif num == 5 and num in pdf_parsed:
            by_key = pdf_parsed[5]["byKey"]
            g702 = pdf_parsed[5]["g702"]
            attachments = pdf_attachments(num) or folder_attachments(num)
        else:
            src = parsed[num]
            by_key = src["byKey"]
            g702 = src["g702"]
            attachments = src["attachments"]

        snapshot = []
        for i, key in enumerate(master_keys):
            # Never inherit final SOV % from master — CO lines added later must start at 0.
            src = by_key.get(key) or prev_by_key.get(key)
            pct = src.get("completedPct", 0) if src else 0
            snapshot.append({"key": key, "completedPct": pct})

        prev_by_key = {s["key"]: {"completedPct": s["completedPct"]} for s in snapshot}

        reqs_out.append(
            {
                "num": num,
                "applicationNumber": f"REQ-{num}",
                "periodTo": g702.get("periodTo", ""),
                "g702": g702,
                "itemsSnapshot": snapshot,
                "attachments": attachments,
            }
        )

    payload = {
        "masterItems": [
            {
                "id": f"item-{i + 1}",
                "section": it["section"],
                "description": it["description"],
                "value": it["value"],
                "contractPct": 0,
                "completedPct": it["completedPct"],
            }
            for i, it in enumerate(master)
        ],
        "contractSum": round(sum(it["value"] for it in master), 2),
        "requisitions": reqs_out,
        "driveFolder": str(DRIVE_BASE),
    }

    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()